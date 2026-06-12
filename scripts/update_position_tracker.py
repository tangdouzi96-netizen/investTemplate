# -*- coding: utf-8 -*-
"""
持仓跟踪数据更新脚本
功能：
1. 从 Excel 读取买入数据（泡泡玛特 09992.HK + 贵州茅台 600519.SH）
2. 通过 akshare + yfinance 获取当前股价和 PE(TTM)，交叉验证
3. 从溯源表读取已报告的年报净利润和 FCF
4. 计算市场隐含增速（Gordon Growth Model 反推）
5. 输出 JSON 到 public/data/position_tracker.json 和 CSV 快照到 data/snapshots/
"""

import json
import os
import re
import sys
from pathlib import Path
from datetime import datetime, timezone, timedelta

# ── Constants ──────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "港股通-泡泡玛特-沪股通-贵州茅台.xlsx"
OUTPUT_JSON = ROOT / "public" / "data" / "position_tracker.json"
SNAPSHOTS_DIR = ROOT / "data" / "snapshots"
TRACE_PREFIX = ROOT / "数据溯源"
TZ_BEIJING = timezone(timedelta(hours=8))
DISCOUNT_RATE = 0.10

# Column mapping for Excel (1-indexed from openpyxl)
COL_EXCHANGE = 1   # A 列：港股通/沪股通
COL_COMPANY = 2    # B 列：公司名
COL_CODE = 3       # C 列：股票代码
COL_BUY_PRICE = 5  # E 列：买入股价
COL_BUY_PE = 6     # F 列：买入 PE
COL_BUY_NP = 8     # H 列：买入净利润
COL_BUY_FCF = 9    # I 列：买入 FCF
COL_BUY_GR = 12    # L 列：买入预期增速
COL_BUY_ER = 14    # N 列：买入预期收益
COL_BUY_LOGIC = 16 # P 列：买入逻辑
COL_MOAT = 17      # Q 列：护城河
COL_RISK = 18      # R 列：风险因素


# ── Helpers ─────────────────────────────────────────────────

def parse_number(val):
    """从混杂字符串/数字中提取 float。例：'150.107港币'→150.107, '127.76亿人民币'→127.76, 823→823.0"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).strip()
    # 移除货币/单位后缀关键词
    val = re.sub(r'[港幣币人民亿元\s]', '', val)
    # 提取第一个数字
    m = re.search(r'[\d,.]+', val)
    if m:
        s = m.group().replace(',', '')
        try:
            return float(s)
        except ValueError:
            pass
    return None


def cell_text(val):
    """返回单元格的清理后文本"""
    if val is None:
        return ""
    return str(val).strip()


# ── Ticker Mapping ─────────────────────────────────────────

def build_ticker(code):
    """根据代码字符串构造 yfinance ticker。
    '09992.HK' -> '09992.HK' (港股)
    '600519.SH' -> '600519.SS' (A股上海)
    '000001.SZ' -> '000001.SZ' (A股深圳)
    """
    if code.endswith(".HK"):
        return code
    if code.endswith(".SS") or code.endswith(".SZ"):
        return code
    if code.endswith(".SH"):
        return code.replace(".SH", ".SS")
    # raw 6-digit code
    if code.startswith("6"):
        return f"{code}.SS"
    else:
        return f"{code}.SZ"


def build_akshare_code(code):
    """根据 yfinance ticker 构造 akshare 的纯数字代码。
    '09992.HK' -> '09992'
    '600519.SS' -> '600519'
    """
    return code.replace(".HK", "").replace(".SS", "").replace(".SZ", "").replace(".SH", "")


# ── B. Excel Reader ────────────────────────────────────────

def read_buy_data(excel_path):
    """读取 Excel 中的买入数据，返回 position 字典列表。"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    positions = []
    # 硬编码映射：公司名→代码、交易所、币种（Excel 中代码列为空）
    hardcoded_codes = {
        "泡泡玛特": {"code": "09992.HK", "exchange": "港股通", "currency": "HKD"},
        "贵州茅台": {"code": "600519.SH", "exchange": "沪股通", "currency": "CNY"},
    }

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]

        company = cell_text(row[COL_COMPANY - 1])
        exchange_raw = cell_text(row[COL_EXCHANGE - 1])

        if not company:
            continue

        # 确定代码和币种
        if company in hardcoded_codes:
            info = hardcoded_codes[company]
            code = info["code"]
            exchange = info["exchange"]
            currency = info["currency"]
        else:
            code = cell_text(row[COL_CODE - 1])
            currency = "HKD" if "HK" in code.upper() else "CNY"

        ticker = build_ticker(code)

        # 提取买入数据
        buy_price = parse_number(row[COL_BUY_PRICE - 1])
        buy_pe = parse_number(row[COL_BUY_PE - 1])
        buy_np = parse_number(row[COL_BUY_NP - 1])
        buy_fcf = parse_number(row[COL_BUY_FCF - 1])
        buy_growth_rate = cell_text(row[COL_BUY_GR - 1])
        buy_expected_return = cell_text(row[COL_BUY_ER - 1])
        buy_logic = cell_text(row[COL_BUY_LOGIC - 1])
        moat = cell_text(row[COL_MOAT - 1])
        risk_factor = cell_text(row[COL_RISK - 1])

        pos = {
            "id": code.split(".")[0],
            "company": company,
            "code": code,
            "ticker": ticker,
            "exchange": exchange,
            "currency": currency,
            "buy": {
                "price": buy_price,
                "pe": buy_pe,
                "net_profit": buy_np,
                "fcf": buy_fcf,
                "growth_rate": buy_growth_rate,
                "expected_return": buy_expected_return,
                "logic": buy_logic,
                "moat": moat,
                "risk_factor": risk_factor,
            },
        }
        positions.append(pos)

    wb.close()
    return positions


# ── C. Price Fetching ──────────────────────────────────────

def fetch_price_akshare(ticker):
    """通过 akshare 获取最新收盘价。"""
    try:
        import akshare as ak
    except ImportError:
        print("[WARN] akshare not installed, skipping akshare price fetch")
        return None

    symbol = build_akshare_code(ticker)
    try:
        if ".HK" in ticker:
            # 港股历史行情
            df = ak.stock_hk_hist(symbol=symbol, period="daily", adjust="qfq",
                                  start_date="20260101", end_date="20501231")
        else:
            # A股历史行情
            df = ak.stock_zh_a_hist(symbol=symbol, period="daily", adjust="qfq",
                                     start_date="20260101")
        if df is not None and not df.empty:
            return float(df["收盘"].iloc[-1])
    except Exception as e:
        print(f"[WARN] akshare price fetch failed for {ticker}: {e}")
    return None


def fetch_price_yfinance(ticker):
    """通过 yfinance 获取最新收盘价。"""
    try:
        import yfinance as yf
    except ImportError:
        print("[WARN] yfinance not installed")
        return None

    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period="5d")
        if not hist.empty:
            return float(hist["Close"].iloc[-1])
    except Exception as e:
        print(f"[WARN] yfinance price fetch failed for {ticker}: {e}")
    return None


# ── D. PE Fetching ─────────────────────────────────────────

def fetch_pe_yfinance(ticker):
    """通过 yfinance info 获取 PE(TTM)。"""
    try:
        import yfinance as yf
    except ImportError:
        return None

    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        pe = info.get("trailingPE")
        if pe is not None and float(pe) > 0 and float(pe) < 10000:
            return float(pe)
    except Exception as e:
        print(f"[WARN] yfinance PE fetch failed for {ticker}: {e}")
    return None


def fetch_pe_akshare(ticker):
    """通过 akshare 获取 PE(TTM)。
    - A股: stock_individual_info_em → 市盈率-动态
    - 港股: stock_hk_financial_indicator_em → 市盈率
    """
    try:
        import akshare as ak
    except ImportError:
        return None

    symbol = build_akshare_code(ticker)

    if ".HK" in ticker:
        # 港股: 东方财富财务指标接口
        try:
            df = ak.stock_hk_financial_indicator_em(symbol=symbol)
            if df is not None and not df.empty and "市盈率" in df.columns:
                pe_val = float(df["市盈率"].iloc[0])
                if pe_val > 0 and pe_val < 10000:
                    return pe_val
        except Exception:
            pass
        return None

    # A股: stock_individual_info_em
    try:
        df = ak.stock_individual_info_em(symbol=symbol)
        if df is not None and not df.empty:
            if hasattr(df, 'columns') and "item" in df.columns:
                row = df[df["item"] == "市盈率-动态"]
                if not row.empty:
                    val = row["value"].iloc[0]
                    if val is not None:
                        return float(val)
    except Exception:
        pass

    return None


def fetch_pe_eastmoney(ticker):
    """备用 PE 获取方法。港股用 financial_indicator_em 已足够，此函数为 A 股备用。"""
    if ".HK" in ticker:
        return None
    # 备用: 可扩展其他 A股 PE 数据源
    return None


# ── E. Cross-Validation ────────────────────────────────────

def cross_validate(v1, v2, threshold=0.01):
    """交叉验证两个数据源的值。
    Returns: (value, is_validated: bool, source: str)
    """
    if v1 is not None and v2 is not None:
        denom = max(abs(v1), abs(v2))
        if denom == 0:
            return 0.0, True, "akshare+yfinance"
        diff_pct = abs(v1 - v2) / denom
        if diff_pct < threshold:
            return round((v1 + v2) / 2, 4), True, "akshare+yfinance"
        else:
            return round((v1 + v2) / 2, 4), False, f"akshare+yfinance(diff={diff_pct:.2%})"
    elif v1 is not None:
        return round(v1, 4), False, "akshare(单源)"
    elif v2 is not None:
        return round(v2, 4), False, "yfinance(单源)"
    else:
        return None, False, "获取失败"


def cross_validate_pe(pe1, pe2, threshold=0.01):
    """交叉验证 PE 数据。"""
    if pe1 is not None and pe2 is not None:
        denom = max(abs(pe1), abs(pe2))
        if denom == 0:
            return 0.0, True, "akshare+eastmoney"
        diff_pct = abs(pe1 - pe2) / denom
        if diff_pct < threshold:
            return round((pe1 + pe2) / 2, 2), True, "akshare+eastmoney"
        else:
            return round((pe1 + pe2) / 2, 2), False, f"akshare+eastmoney(diff={diff_pct:.2%})"
    elif pe1 is not None:
        return round(pe1, 2), False, "akshare(单源)"
    elif pe2 is not None:
        return round(pe2, 2), False, "yfinance(单源)"
    else:
        return None, False, "获取失败"


def fetch_with_validation(ticker):
    """获取股价和 PE，进行交叉验证。
    Returns: dict with price{}, pe_ttm{}, and validation flags.
    """
    # 价格：akshare + yfinance
    p_ak = fetch_price_akshare(ticker)
    p_yf = fetch_price_yfinance(ticker)
    p_val, p_ok, p_src = cross_validate(p_ak, p_yf)

    # PE：akshare + yfinance (trailingPE from info)
    pe_ak = fetch_pe_akshare(ticker)
    pe_yf = fetch_pe_yfinance(ticker)
    pe_val, pe_ok, pe_src = cross_validate_pe(pe_ak, pe_yf)

    result = {
        "price": {"value": p_val, "source": p_src, "cross_validated": p_ok},
        "pe_ttm": {"value": pe_val, "source": pe_src, "cross_validated": pe_ok},
        "p_validation": p_ok,
        "pe_validation": pe_ok,
    }
    return result


# ── F. Traceability Table Reader ───────────────────────────

def find_trace_file(company, code):
    """查找对应的溯源表文件。"""
    code_stripped = code.split(".")[0]
    candidates = list(TRACE_PREFIX.glob(f"*{company}*{code_stripped}*.md"))
    if not candidates:
        candidates = list(TRACE_PREFIX.glob(f"*{code_stripped}*.md"))
    if not candidates:
        candidates = list(TRACE_PREFIX.glob(f"*{company}*.md"))
    return candidates[0] if candidates else None


def read_traceability_reported(company, code):
    """从溯源表 markdown 读取最近一年的归母净利润和 FCF。
    返回 (net_profit: float or None, fcf: float or None, year: str or None)
    """
    filepath = find_trace_file(company, code)
    if filepath is None:
        print(f"[WARN] No traceability file found for {company} ({code})")
        return None, None, None

    content = filepath.read_text(encoding="utf-8")
    net_profit = None
    fcf = None
    year = None

    def _is_formula(text):
        """判断单元格内容看起来像公式而非纯数字。"""
        return bool(re.search(r'[+/\-−×x\*]', text))

    def _clean_cell(text):
        """清理单元格：移除 ** 标记、✅ 等标记、页码引用 (pN)。"""
        cleaned = re.sub(r'\*+|✅|📝|🔗|➖|⚠️|📄|🟡|🔴|⚙️', '', text)
        # 移除页码引用: (p54), (p61) 等
        cleaned = re.sub(r'\s*\(p\d+[^)]*\)', '', cleaned)
        return cleaned.strip()

    # ── 解析归母净利润 (A06 for 泡泡玛特, A11 for 贵州茅台) ──
    # 查找 A 区 PL 表中的"归母净利润"行（限 A0x 或 A1x 开头的行）
    for line in content.split("\n"):
        # 仅匹配 A 区编号行（如 | A06 | ... 或 | A11 | ...）
        if re.search(r'\|\s*A\d{1,2}\s*\|.*归[母属]净利[润]', line):
            cells = [c.strip() for c in line.split("|")]
            cells = [c for c in cells if c]  # 去首尾空
            if len(cells) >= 3:
                # 从右往左找第一个"看起来像有效财务数字"的纯数字
                for cell_str in reversed(cells[2:]):
                    stripped = _clean_cell(cell_str)
                    if _is_formula(stripped) or '%' in stripped or stripped in ('--', '-', ''):
                        continue
                    # 尝试匹配纯数字（可能带小数点）
                    m = re.match(r'^([\d,.]+)\s*$', stripped)
                    if m:
                        val = float(m.group(1).replace(",", ""))
                        if val > 0 and val < 1000000:
                            net_profit = val
                            break
            if net_profit is not None:
                break

    # ── 解析 FCF ──
    # 策略：优先匹配 B 区 B01 或 B03 行（单行展示），取最近的年份值
    # 泡泡玛特 B03: | B03 | FCF（CAPEX=PPE+无形） | ... | **96.94** | 44.37 |
    # 茅台 B01:     | B01 | FCF（自由现金流）     | 606.20 | 313.92 | ...
    # 需要跳过 B09 (FCF/归母净利) 和 B17 (5年趋势) 这种行

    for line in content.split("\n"):
        # 只匹配 B01 或 B03（单行计算指标，非趋势行 B17）
        if re.search(r'\|\s*B0[13]\s*\|.*FCF', line):
            cells = [c.strip() for c in line.split("|")]
            cells = [c for c in cells if c]
            if len(cells) >= 3:
                # 收集所有纯数字（从B区单元格）
                numbers = []
                for cell_str in cells[2:]:
                    stripped = _clean_cell(cell_str)
                    if _is_formula(stripped) or '%' in stripped or stripped in ('--', '-', ''):
                        continue
                    m = re.match(r'^([\d,.]+)\s*$', stripped)
                    if m:
                        val = float(m.group(1).replace(",", ""))
                        if val > 0 and val < 100000:
                            numbers.append(val)
                if numbers:
                    # B01 行（茅台）：5个数字，按年份从左到右排列 → 取最后一个
                    # B03 行（泡泡玛特）：2-3个数字，2025结果在前，2024结果在后 → 取第一个
                    if len(numbers) >= 5:
                        fcf = numbers[-1]  # 茅台：5年趋势取最后一年
                    elif len(numbers) == 2:
                        fcf = numbers[0]  # 泡泡玛特：近年在前面
                    else:
                        # 不确定时取最大的（最近年份通常是最大值）
                        fcf = max(numbers)
            if fcf is not None:
                break

    # 如果 B01/B03 没找到，尝试 B17 (5年趋势行)
    if fcf is None:
        for line in content.split("\n"):
            if re.search(r'\|\s*B17\s*\|.*FCF', line):
                cells = [c.strip() for c in line.split("|")]
                cells = [c for c in cells if c]
                if len(cells) >= 3:
                    numbers = []
                    for cell_str in cells[2:]:
                        stripped = _clean_cell(cell_str)
                        if _is_formula(stripped) or '%' in stripped or stripped in ('--', '-', ''):
                            continue
                        m = re.match(r'^([\d,.]+)\s*$', stripped)
                        if m:
                            val = float(m.group(1).replace(",", ""))
                            if val > 0:
                                numbers.append(val)
                    if numbers:
                        fcf = numbers[-1]
                break

    # ── 推导年份 ──
    fy_m = re.search(r'FY(20\d{2})', content)
    if fy_m:
        year = f"FY{fy_m.group(1)}"
    else:
        years = re.findall(r'(20\d{2})[年_末]', content)
        if years:
            year = f"FY{max(int(y) for y in years)}"
        else:
            year = "FY2025"

    return net_profit, fcf, year


# Hardcoded fallback as last resort
TRACE_FALLBACK = {
    "泡泡玛特": {"net_profit": 127.76, "fcf": 96.94, "year": "FY2025"},
    "贵州茅台": {"net_profit": 823.20, "fcf": 583.94, "year": "FY2025"},
}


def read_traceability_buy_data(company, code):
    """读取溯源表数据，解析失败时用硬编码 fallback。"""
    np_val, fcf_val, year = read_traceability_reported(company, code)
    used_fallback = False
    if np_val is None and company in TRACE_FALLBACK:
        fb = TRACE_FALLBACK[company]
        np_val = fb["net_profit"]
        fcf_val = fb["fcf"]
        year = fb["year"]
        used_fallback = True
    if used_fallback:
        print(f"  [INFO] Fallback trace data for {company}: NP={np_val}, FCF={fcf_val}, {year}")
    return np_val, fcf_val, year


# ── G. Market Implied Calculation ──────────────────────────

def compute_market_implied(pe_ttm):
    """Gordon Growth Model 反推市场隐含永续增速。
    P = E / (r - g) → earnings_yield = 1/PE = r - g → g = r - 1/PE
    其中 r = DISCOUNT_RATE (10%)

    Returns: dict with implied_growth (%), implied_return (%), and notes.
    """
    result = {
        "implied_growth": None,
        "implied_return": None,
        "growth_note": "当前PE反推市场隐含永续增速，假设折现率10%",
        "return_note": "隐含增速 -> 预期年化收益（含分红）",
    }

    if pe_ttm is None or pe_ttm <= 0:
        return result

    earnings_yield = 1.0 / pe_ttm
    implied_growth = max(0, DISCOUNT_RATE - earnings_yield)
    implied_return = DISCOUNT_RATE  # 市场均衡下定价于折现率

    result["implied_growth"] = round(implied_growth * 100, 2)
    result["implied_return"] = round(implied_return * 100, 2)
    return result


# ── H. Main Pipeline ───────────────────────────────────────

def run_update():
    """运行完整的更新流程。"""
    print("=" * 60)
    print("Position Tracker Update")
    print(f"Time: {datetime.now(TZ_BEIJING).isoformat()}")
    print("=" * 60)

    # 1. 读取买入数据
    print("\n[1/5] Reading buy data from Excel...")
    positions = read_buy_data(EXCEL_PATH)
    if not positions:
        print("[ERROR] No positions found in Excel file")
        return 1
    print(f"  Found {len(positions)} positions")
    for p in positions:
        print(f"    - {p['company']} ({p['code']}): buy_price={p['buy']['price']}, "
              f"buy_pe={p['buy']['pe']}, ticker={p['ticker']}")

    # 2. 读取旧 JSON（保留 subjective_return）
    old_subjective = {}
    if OUTPUT_JSON.exists():
        try:
            with open(OUTPUT_JSON, "r", encoding="utf-8") as f:
                old_data = json.load(f)
            for op in old_data.get("positions", []):
                sr = op.get("current", {}).get("subjective_return", "")
                if sr:
                    old_subjective[op["id"]] = sr
            print(f"\n  Loaded {len(old_subjective)} old subjective_return values")
        except Exception as e:
            print(f"\n  [WARN] Could not read old JSON: {e}")

    # 3. 获取行情和财务数据
    print("\n[2/5] Fetching live prices and PE...")
    validation_stats = {"total_checks": 0, "passed": 0, "failed": 0, "single_source": 0}

    for pos in positions:
        ticker = pos["ticker"]
        company = pos["company"]
        code = pos["code"]

        print(f"\n  --- {company} ({code}) ---")

        # a. 获取价格和 PE
        print(f"    Fetching via akshare + yfinance (ticker={ticker})...")
        fv = fetch_with_validation(ticker)

        price_val = fv["price"]["value"]
        pe_val = fv["pe_ttm"]["value"]

        if price_val is not None:
            print(f"    Price: {price_val:.2f} (validated={'Y' if fv['p_validation'] else 'N'}, "
                  f"src={fv['price']['source']})")
        else:
            print(f"    Price: FAILED")

        if pe_val is not None:
            print(f"    PE:    {pe_val:.2f} (validated={'Y' if fv['pe_validation'] else 'N'}, "
                  f"src={fv['pe_ttm']['source']})")
        else:
            print(f"    PE:    FAILED")

        # 更新验证统计
        for vflag in [fv["p_validation"], fv["pe_validation"]]:
            validation_stats["total_checks"] += 1
            if vflag:
                validation_stats["passed"] += 1
            else:
                validation_stats["failed"] += 1
        # 另外统计单源情况
        for src_str in [fv["price"]["source"], fv["pe_ttm"]["source"]]:
            if "单源" in src_str:
                validation_stats["single_source"] += 1

        # b. 读取溯源表数据
        print(f"    Reading traceability data...")
        np_reported, fcf_reported, year = read_traceability_buy_data(company, code)
        if np_reported is not None:
            print(f"    Net Profit (reported): {np_reported} 亿CNY ({year})")
        if fcf_reported is not None:
            print(f"    FCF (reported):        {fcf_reported} 亿CNY ({year})")

        # c. 计算市场隐含
        implied = compute_market_implied(pe_val)
        if implied["implied_growth"] is not None:
            print(f"    Market Implied Growth: {implied['implied_growth']:.2f}%")
        else:
            print(f"    Market Implied Growth: N/A (PE unavailable)")

        # d. 组装 current 块
        pos_id = pos["id"]
        sr = old_subjective.get(pos_id, pos["buy"]["expected_return"])

        pos["current"] = {
            "price": {
                "value": fv["price"]["value"],
                "source": fv["price"]["source"],
                "cross_validated": fv["p_validation"],
            },
            "pe_ttm": {
                "value": fv["pe_ttm"]["value"],
                "source": fv["pe_ttm"]["source"],
                "cross_validated": fv["pe_validation"],
            },
            "net_profit_reported": {
                "value": np_reported,
                "unit": "亿CNY",
                "year": year if year else "FY2025",
                "source": "溯源表-B区",
            },
            "net_profit_consensus": {
                "value": None,
                "unit": "亿CNY",
                "year": "FY2026E",
                "source": "东方财富一致预期",
            },
            "fcf_reported": {
                "value": fcf_reported,
                "unit": "亿CNY",
                "year": year if year else "FY2025",
                "source": "溯源表-B区",
            },
            "fcf_consensus": {
                "value": None,
                "unit": "亿CNY",
                "year": "FY2026E",
                "source": "东方财富一致预期",
            },
            "growth_consensus": {
                "value": None,
                "source": "",
            },
            "market_implied": implied,
            "subjective_return": sr,
        }

    # 4. 写入 JSON
    print("\n[3/5] Writing JSON...")
    now_iso = datetime.now(TZ_BEIJING).isoformat()

    output = {
        "meta": {
            "updated_at": now_iso,
            "version": "1.0",
            "validation": validation_stats,
        },
        "positions": [
            {
                "id": p["id"],
                "company": p["company"],
                "code": p["code"],
                "exchange": p["exchange"],
                "currency": p["currency"],
                "buy": p["buy"],
                "current": p["current"],
            }
            for p in positions
        ],
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  JSON written: {OUTPUT_JSON}")
    print(f"  Size: {OUTPUT_JSON.stat().st_size / 1024:.1f} KB")

    # 5. 写入 CSV 快照
    print("\n[4/5] Writing CSV snapshot...")
    today_str = datetime.now(TZ_BEIJING).strftime("%Y-%m-%d")
    csv_path = SNAPSHOTS_DIR / f"{today_str}.csv"
    SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)

    csv_lines = []
    csv_lines.append("日期,公司,代码,当前股价,买入股价,盈亏%,当前PE,买入PE,"
                     "年报净利润,买入净利润,年报FCF,买入FCF,"
                     "市场隐含增速%,主观预期收益,验证股价,验证PE")

    for p in positions:
        cur = p["current"]
        buy = p["buy"]

        price_cur = cur["price"]["value"]
        price_buy = buy["price"]
        pnl_pct = ""
        if price_cur is not None and price_buy is not None and price_buy > 0:
            pnl_pct = f"{(price_cur / price_buy - 1) * 100:.2f}"

        pe_cur = cur["pe_ttm"]["value"]
        pe_buy = buy["pe"]

        np_r = cur["net_profit_reported"]["value"]
        np_b = buy["net_profit"]
        fcf_r = cur["fcf_reported"]["value"]
        fcf_b = buy["fcf"]

        ig = cur["market_implied"]["implied_growth"]
        ig_str = f"{ig:.2f}" if ig is not None else ""

        sr_str = cur["subjective_return"].replace(",", " ") if cur["subjective_return"] else ""

        vp = "Y" if cur["price"]["cross_validated"] else "N"
        vpe = "Y" if cur["pe_ttm"]["cross_validated"] else "N"

        line = (f"{today_str},{p['company']},{p['code']},"
                f"{price_cur if price_cur else ''},{price_buy if price_buy else ''},"
                f"{pnl_pct},"
                f"{pe_cur if pe_cur else ''},{pe_buy if pe_buy else ''},"
                f"{np_r if np_r else ''},{np_b if np_b else ''},"
                f"{fcf_r if fcf_r else ''},{fcf_b if fcf_b else ''},"
                f"{ig_str},{sr_str},{vp},{vpe}")
        csv_lines.append(line)

    with open(csv_path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(csv_lines) + "\n")
    print(f"  CSV written: {csv_path}")

    # 6. 打印摘要
    print("\n[5/5] Summary")
    print(f"  Positions: {len(positions)}")
    print(f"  Validation: {validation_stats['passed']}/{validation_stats['total_checks']} passed, "
          f"{validation_stats['failed']} failed, {validation_stats['single_source']} single-source")

    for p in positions:
        cur = p["current"]
        price = cur["price"]["value"]
        pe = cur["pe_ttm"]["value"]
        ig = cur["market_implied"]["implied_growth"]
        price_str = f"{price:.2f}" if price is not None else "N/A"
        pe_str = f"{pe:.2f}" if pe is not None else "N/A"
        ig_str = f"{ig:.2f}%" if ig is not None else "N/A"
        print(f"  {p['company']}: price={price_str}, PE={pe_str}, implied_growth={ig_str}")

    print("\n[DONE] Position tracker update completed")
    return 0


# ── Main ────────────────────────────────────────────────────

if __name__ == "__main__":
    sys.exit(run_update())
